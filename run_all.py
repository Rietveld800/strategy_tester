# run_all.py
#
# Run a strategy (see strategies.py) over EVERY market's daily array files and write one
# xlsx workbook per strategy. Obsolete markets (data collection stopped) are included so
# their trades can be reviewed too.
#
# Each market is an INDEPENDENT backtest starting from a fresh STARTING_CAPITAL, so the
# per-market returns are directly comparable. The window is data-driven: trading begins
# once a market's reversals are first reported (older files carry no reversal block) and
# runs to its last daily bar. The shared-account view is run_portfolio.py.
#
#   python run_all.py            # every registered strategy
#   python run_all.py slowfix    # just one
#
# Output: output/<strategy>_all_markets_daily.xlsx

import sys
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import engine as eng
import strategies

TRADE_COLS = ["market", "id", "side", "entry_date", "exit_date", "bars_in_trade",
              "entry", "stop", "target", "exit_price", "exit_reason",
              "r_multiple", "pnl_pct", "equity_before", "equity_after"]
SUMMARY_COLS = ["market", "obsolete", "first_reversal", "last_bar", "bars", "trades",
                "wins", "stops", "unknown", "data_end", "open_at_end", "win_rate_%",
                "return_%", "final_equity"]


def out_path(strategy):
    return eng.OUT_DIR / f"{strategy.key}_all_markets_daily.xlsx"


def first_reversal_date(bars):
    """First bar carrying BOTH bullish and bearish reversals -- the strategy's real start."""
    for b in bars:
        if b.bull and b.bear:
            return b.date
    return None


def write(strategy, results):
    """One workbook for `strategy` from a precomputed engine.run_markets() result list."""
    newest = max(m["bars"][-1].date for m in results)
    per_market = []
    for m in results:
        # `obsolete` comes from run_markets, which already had to decide it before the
        # backtest ran (an obsolete market's open trade is flattened at its last close).
        # Recomputing it here would be a second definition waiting to drift.
        per_market.append(dict(
            name=m["name"], bars=m["bars"], res=m["res"][strategy.key],
            first_rev=first_reversal_date(m["bars"]), last_bar=m["bars"][-1].date,
            obsolete=m["obsolete"]))

    wb = Workbook()
    _sheet_summary(wb.active, strategy, per_market, newest)
    _sheet_trades(wb.create_sheet("trades"), per_market)
    path = out_path(strategy)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"written: {path}")


def main(argv):
    picked = strategies.selected(argv)
    # Only the strategies' own Rule 4 settings -- this workbook is written at the default,
    # and the report's cap dial never changes a file on disk. The caps are named explicitly;
    # a strategy outside the cap family is added by run_markets from its own token, so it
    # must NOT be listed here as a cap (its `cap` is None, which means "no cap" and is a
    # different run entirely).
    results = eng.run_markets(picked, caps=[s.cap for s in picked if s.r4.in_grid])
    print()
    for s in picked:
        write(s, results)


def _summ_row(m):
    trades = m["res"]["trades"]
    closed = [t for t in trades if t["exit_reason"] != "open_at_end"]
    wins = sum(1 for t in closed if t["pnl_pct"] > 0)
    stops = sum(1 for t in closed if t["exit_reason"] == "stop")
    unknown = sum(1 for t in closed if t["exit_reason"] == "unknown_pl")
    data_end = sum(1 for t in closed if t["exit_reason"] == "data_end")
    open_end = sum(1 for t in trades if t["exit_reason"] == "open_at_end")
    ret = (m["res"]["final_equity"] / eng.STARTING_CAPITAL - 1) * 100
    win_rate = (100 * wins / len(closed)) if closed else None
    return dict(market=m["name"], obsolete="yes" if m["obsolete"] else "",
                first_reversal=str(m["first_rev"].date()) if m["first_rev"] else "",
                last_bar=str(m["last_bar"].date()), bars=len(m["bars"]),
                trades=len(closed), wins=wins, stops=stops, unknown=unknown,
                data_end=data_end, open_at_end=open_end,
                win_rate=round(win_rate, 1) if win_rate is not None else None,
                ret=round(ret, 2), final_equity=round(m["res"]["final_equity"], 2))


HEAD_FILL = PatternFill("solid", fgColor="1F3864")
HEAD_FONT = Font(bold=True, color="FFFFFF")
OBS_FILL = PatternFill("solid", fgColor="EEEEEE")
POS_FONT = Font(color="1B7A34")
NEG_FONT = Font(color="B02418")


def _header(ws, cols):
    for c, name in enumerate(cols, 1):
        cell = ws.cell(1, c, name)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"


def _autosize(ws, cols):
    for c, name in enumerate(cols, 1):
        width = max(len(str(name)) + 2,
                    *(len(str(ws.cell(r, c).value or "")) + 2
                      for r in range(2, ws.max_row + 1)) or [0])
        ws.column_dimensions[get_column_letter(c)].width = min(width, 42)


def _sheet_summary(ws, strategy, per_market, newest):
    ws.title = "summary"
    _header(ws, SUMMARY_COLS)
    # active markets first (by return desc), then obsolete (by return desc) -- charter's order
    rows = [_summ_row(m) for m in per_market]
    rows.sort(key=lambda r: (r["obsolete"] == "yes", -r["ret"]))
    for r in rows:
        vals = [r["market"], r["obsolete"], r["first_reversal"], r["last_bar"], r["bars"],
                r["trades"], r["wins"], r["stops"], r["unknown"], r["data_end"],
                r["open_at_end"], r["win_rate"], r["ret"], r["final_equity"]]
        ws.append(vals)
        row = ws.max_row
        ret_cell = ws.cell(row, SUMMARY_COLS.index("return_%") + 1)
        ret_cell.number_format = "+0.00;-0.00"
        ret_cell.font = POS_FONT if r["ret"] > 0 else (NEG_FONT if r["ret"] < 0 else Font())
        ws.cell(row, SUMMARY_COLS.index("final_equity") + 1).number_format = "#,##0"
        if r["obsolete"] == "yes":
            for c in range(1, len(SUMMARY_COLS) + 1):
                ws.cell(row, c).fill = OBS_FILL
    _autosize(ws, SUMMARY_COLS)
    # footer note
    note = (f"{strategy.key} daily ({strategy.rule4}), "
            f"starting capital {eng.STARTING_CAPITAL:,.0f}, "
            f"risk {eng.RISK_PCT}%/trade, fees {eng.FEES}. "
            f"obsolete = last bar > {eng.OBSOLETE_AFTER_DAYS}d before newest "
            f"({newest.date()}); a position still open on an obsolete market's last bar is "
            f"flattened at that bar's close (data_end). "
            f"generated {datetime.now(timezone.utc):%Y-%m-%d %H:%M UTC}.")
    ws.append([])
    ws.append([note])


def _sheet_trades(ws, per_market):
    _header(ws, TRADE_COLS)
    order = sorted(per_market, key=lambda m: m["name"])
    for m in order:
        for t in m["res"]["trades"]:
            ws.append([m["name"], t["id"], t["side"], t["entry_date"], t["exit_date"],
                       t["bars_in_trade"], t["entry"], t["stop"], t["target"],
                       t["exit_price"], t["exit_reason"], t["r_multiple"], t["pnl_pct"],
                       t["equity_before"], t["equity_after"]])
            row = ws.max_row
            pnl = t["pnl_pct"]
            if pnl is not None:
                cell = ws.cell(row, TRADE_COLS.index("pnl_pct") + 1)
                cell.number_format = "+0.00;-0.00"
                cell.font = POS_FONT if pnl > 0 else (NEG_FONT if pnl < 0 else Font())
    _autosize(ws, TRADE_COLS)


if __name__ == "__main__":
    main(sys.argv[1:])
