# run_portfolio.py
#
# Combine every market's trades into ONE shared account and build a chronological equity
# curve of the capital, per strategy. The per-market trades themselves are unchanged
# (entries, exits and R-multiples depend only on price/reversals, not capital); only the
# money management is re-run on the shared account.
#
#   python run_portfolio.py            # every registered strategy
#   python run_portfolio.py quickfixwick   # just one
#
# Money management (confirmed with the user):
#   - one account, starting STARTING_CAPITAL; cash balance changes ONLY when a trade closes.
#   - a new trade risks RISK_PCT of LIQUID capital = cash balance - risk already tied up in
#     the trades currently open. Each open trade ties up its own risk until it closes.
#   - no cap on how many trades are open at once.
#   - within a date: process ALL entries first (sized while that day's closing trades still
#     tie up their risk), THEN the exits (which book P&L and grow the balance).
#   - several entries on one date are sized in a fixed order (alphabetical by market), each
#     off the base left by the earlier ones.
#
# An "open at end" position (data ended while it was open) ties up its risk from entry until
# its market's last bar, then releases it with zero realised P&L -- its outcome is unknown.
#
# Outputs: output/<strategy>_portfolio_daily.xlsx  and  output/_equity_<strategy>.json
# (the latter feeds build_equity_html.py).

import json
import sys
from collections import defaultdict
from datetime import date

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import engine as eng
import strategies

# The risk the account is currently being run at. A MODULE-LEVEL dial rather than an argument
# threaded through every function: `account`, `_replay` and the whole workbook are one
# simulation, and passing the same number through five layers only invites one of them being
# missed. `at_risk()` sets it for the duration of one run and always puts it back, so a caller
# can never leak a strategy's risk into the next one. Defaults to the reference risk, which is
# what the shared variant grid is priced at.
RISK_PCT = eng.RISK_PCT
START_CAP = eng.STARTING_CAPITAL


class at_risk:
    """`with at_risk(0.457):` -- run the money management at that risk, then restore."""

    def __init__(self, risk):
        self.risk = risk

    def __enter__(self):
        global RISK_PCT
        self.prev, RISK_PCT = RISK_PCT, self.risk
        return self

    def __exit__(self, *exc):
        global RISK_PCT
        RISK_PCT = self.prev
        return False

# --- transaction costs ------------------------------------------------------------
# Realistic fills, charged as SLIPPAGE in ticks (the market's inferred price tick, the
# same unit the stop uses). Entry and the limit take-profit fill at/near their price; a
# stop is a market order on trigger and slips more, especially on gaps. Cost converts to
# R via each trade's own risk distance -- ticks * tick / risk_per_unit -- so a wide-risk
# trade barely feels it and a tight-risk trade feels it more, which is how slippage really
# behaves. In dollars this equals ticks * tick * position size, the true slippage cost.
SLIP_ENTRY_TICKS = 1       # entry fill one tick worse than the reversal price
SLIP_TARGET_TICKS = 1      # limit take-profit / reversal exit
SLIP_STOP_TICKS = 3        # stop-loss and unknown-outcome exits slip more (market/gap)


def out_xlsx(strategy):
    return eng.OUT_DIR / f"{strategy.key}_portfolio_daily.xlsx"


def data_json(strategy):
    return eng.OUT_DIR / f"_equity_{strategy.key}.json"


def cost_in_r(trade):
    """Round-trip slippage for one trade, expressed in R (a fraction of the risk taken)."""
    rpu = trade.get("risk_per_unit")
    if not rpu or rpu <= 0:
        return 0.0
    reason = trade["exit_reason"]
    if reason == "open_at_end":
        exit_ticks = 0                                   # never exited -> no exit slippage
    elif reason in ("stop", "unknown_pl"):
        exit_ticks = SLIP_STOP_TICKS
    else:
        # Includes 'data_end' (an obsolete market flattened at its last close): that is an
        # orderly close-out at the bell, not a stop triggered into a gap, so it is charged
        # the limit/target rate rather than the stop rate.
        #
        # And it includes the BAR EXITS, for the same reason. 'exit_close' is a
        # market-on-close and 'exit_open' a market-on-open: both are scheduled orders placed
        # into the most liquid minutes of the session, so they belong with the limit rate
        # rather than with a stop that fires into a move nobody chose the timing of. They
        # are the whole strategy for the 24 quick exits, so the choice matters more there
        # than anywhere else -- at 2 ticks round trip it is already the difference between a
        # winner and a loser on quickfixclose0's smallest trades. (A bar-exit strategy that is
        # STOPPED out is charged the stop rate like anything else; only the scheduled exit
        # gets the limit rate.)
        exit_ticks = SLIP_TARGET_TICKS
    return (SLIP_ENTRY_TICKS + exit_ticks) * trade["tick"] / rpu


def collect(results, trades_of):
    """Trades tagged with market + last bar, plus the set of all trading days.

    `trades_of(m)` picks which backtest of a market to take -- a strategy's default run for
    the on-disk outputs, one cap of the grid for the report's variants.
    """
    raw, all_days = [], set()
    for m in results:
        last_bar = m["bars"][-1].date.date()
        all_days.update(b.date.date() for b in m["bars"])
        for t in trades_of(m):
            raw.append({**t, "market": m["name"], "last_bar": last_bar, "tick": m["tick"]})
    return raw, sorted(all_days)


def prepare(raw):
    """Normalise each trade to (entry day, exit day, gross R, cost, net R), in place.

    An open-at-end trade "exits" on its market's last bar with gross R = 0 (outcome
    unknown): it ties up its risk until then and releases it having booked nothing.
    """
    for t in raw:
        t["entry_d"] = date.fromisoformat(t["entry_date"])
        if t["exit_reason"] == "open_at_end":
            t["exit_d"], t["gross_r"] = t["last_bar"], 0.0
        else:
            t["exit_d"], t["gross_r"] = date.fromisoformat(t["exit_date"]), t["r_multiple"]
        t["cost_r"] = cost_in_r(t)
        t["r"] = t["gross_r"] - t["cost_r"]          # net R actually booked
    return raw


def first_trading_day(raw, all_days):
    """The day the shared account starts working: its first entry."""
    return min((t["entry_d"] for t in raw), default=(all_days[0] if all_days else None))


def account(raw, all_days, first_day):
    """The shared-account money management over `raw` (already prepared).

    Returns (timeline, stats) and fills each trade's risk_dollars / base_at_entry /
    pnl_dollars / balance_after. Split out of `run` so the cap grid can replay the exact
    same loop for every cap without a second copy of the rules.
    """
    entries_by_day, exits_by_day = defaultdict(list), defaultdict(list)
    for i, t in enumerate(raw):
        entries_by_day[t["entry_d"]].append(i)
        exits_by_day[t["exit_d"]].append(i)

    cash, committed = START_CAP, 0.0
    open_risk = {}                         # trade index -> risk dollars tied up
    timeline, peak, max_dd = [], START_CAP, 0.0
    max_open, max_committed_ratio, total_cost, days_in_mkt = 0, 0.0, 0.0, 0

    for day in all_days:
        if first_day and day < first_day:
            continue
        # 1) ENTRIES first -- sized while that day's closing trades still tie up their risk
        ent_notes = []
        for i in sorted(entries_by_day.get(day, []), key=lambda j: raw[j]["market"]):
            base = max(cash - committed, 0.0)        # guard: never size on negative capital
            risk_d = base * RISK_PCT / 100.0
            raw[i]["risk_dollars"] = risk_d
            raw[i]["base_at_entry"] = base
            open_risk[i] = risk_d
            committed += risk_d
            ent_notes.append(f"{raw[i]['market']} {raw[i]['side']} (risk {risk_d:,.0f})")

        # 2) EXITS -- book P&L and release the tied-up risk
        ex_notes = []
        for i in exits_by_day.get(day, []):
            risk_d = open_risk.pop(i)
            committed -= risk_d
            pnl = raw[i]["r"] * risk_d
            cash += pnl
            total_cost += raw[i]["cost_r"] * risk_d       # slippage paid, in dollars
            raw[i]["pnl_dollars"] = pnl
            raw[i]["balance_after"] = cash
            tag = "open->closed 0" if raw[i]["exit_reason"] == "open_at_end" else f"{pnl:+,.0f}"
            ex_notes.append(f"{raw[i]['market']} {raw[i]['exit_reason']} ({tag})")

        liquid = cash - committed
        peak = max(peak, cash)
        dd = (cash - peak) / peak * 100.0
        max_dd = max(max_dd, -dd)
        max_open = max(max_open, len(open_risk))
        if len(open_risk) > 0:
            days_in_mkt += 1
        if cash > 0:
            max_committed_ratio = max(max_committed_ratio, committed / cash * 100.0)
        timeline.append(dict(
            date=day, cash=cash, committed=committed, liquid=liquid, dd=dd,
            open_count=len(open_risk),
            open_markets=", ".join(sorted(raw[i]["market"] for i in open_risk)),
            entries="; ".join(ent_notes), exits="; ".join(ex_notes)))

    gross_final = _replay(raw, all_days, first_day, "gross_r")   # same trades, no costs
    closed = [t for t in raw if t["exit_reason"] != "open_at_end"]
    wins = sum(1 for t in closed if t["r"] > 0)
    n_days = len(timeline)
    stats = dict(final=cash, ret=(cash / START_CAP - 1) * 100.0, max_dd=max_dd,
                 n_trades=len(closed), wins=wins,
                 win_rate=100 * wins / len(closed) if closed else 0.0,
                 max_open=max_open, max_committed_pct=max_committed_ratio,
                 first=first_day, last=all_days[-1] if all_days else None,
                 n_markets=len({t["market"] for t in raw}),
                 gross_final=gross_final, gross_ret=(gross_final / START_CAP - 1) * 100.0,
                 total_cost=total_cost, avg_bars=_avg_bars(closed),
                 time_in_market=100 * days_in_mkt / n_days if n_days else 0.0)
    stats.update(_streaks_and_avgs(raw))
    return timeline, stats


def run(strategy, results):
    """One strategy's on-disk outputs, at its DEFAULT Rule 4: the workbook and the JSON."""
    raw, all_days = collect(results, lambda m: m["res"][strategy.key]["trades"])
    # EVERY market that was backtested, not just the ones that produced a trade. A market
    # the rules never fired on is a real result and belongs in the per-market table as a
    # zero, otherwise the report reads as though only 28 markets were ever researched.
    universe = [dict(m=m["name"], obs=bool(m["obsolete"])) for m in results]

    prepare(raw)
    first_day = first_trading_day(raw, all_days)
    # This strategy's OWN risk -- the one that puts it at the 6% drawdown budget. Every
    # figure in the workbook and on the page is at that risk, so all three strategies are
    # published at equal pain rather than at equal bet size.
    with at_risk(strategy.risk_pct):
        timeline, stats = account(raw, all_days, first_day)
    # The same account replayed at the REFERENCE risk, purely so build_equity_html.py can
    # check the shared variant grid against this workbook. The grid is priced at the
    # reference risk (it is shared by every page), so the guard has to compare like with
    # like; without this it would flag every strategy whose own risk is not the reference.
    with at_risk(eng.RISK_PCT):
        ref_final = account(raw, all_days, first_day)[1]["final"]
    stats["ref_final"] = ref_final
    stats["n_markets_all"] = len(universe)
    _write(strategy, raw, timeline, stats)
    _write_json(strategy, raw, timeline, stats, universe)
    print(f"{strategy.key} (rule 4: {strategy.r4.label}, risk {strategy.risk_pct:g}%): "
          f"NET  ${stats['final']:,.2f} ({stats['ret']:+.2f}%)   "
          f"gross ${stats['gross_final']:,.0f} ({stats['gross_ret']:+.2f}%)   "
          f"cost drag ${stats['total_cost']:,.0f}")
    print(f"  maxDD {stats['max_dd']:.2f}%  trades {stats['n_trades']} "
          f"winrate {stats['win_rate']:.1f}%  "
          f"max concurrent {stats['max_open']}  "
          f"time in market {stats['time_in_market']:.0f}%  "
          f"avg hold {stats['avg_bars']:.1f}d")
    print(f"  written: {out_xlsx(strategy)}")


# --- the Rule 4 variant grid, for the report's dial --------------------------------
# One packed table of EVERY cap in strategies.CAP_CHOICES, written once and shared by every
# strategy's page -- the cap family is a single family, so two strategies sitting on the
# same cap are one run and must not be stored twice. Every registered Rule 4 that is NOT a cap
# (the wick target and the 23 bar exits) is filed in the same table under its own
# token; they are listed separately from `caps` in the document, because the caps are the
# dial's axis and a different Rule 4 shape is not a point on it.
#
# Why precomputed at all, when the risk dial replays in the browser: risk only changes the
# DOLLAR SIZING of a fixed trade list, but Rule 4 changes the trades themselves. It moves
# every exit, and because only one position per market runs at a time, an earlier exit frees
# that market for a later signal a longer hold would have missed. There is nothing to replay
# from -- each setting is a real backtest, so all of them are run here.
VARIANTS_PATH = eng.OUT_DIR / "_variants.json"

# The packed row layout. Kept as data because build_equity_html.py's JS unpacks by this
# exact order -- change one side and you must change the other.
VAR_COLS = ["m", "side", "din", "dout", "xd", "gr", "cr", "bars",
            "pin", "pout", "sl", "reason"]
# APPEND-ONLY. The table travels WITH the rows, so the order is technically free to change,
# but keeping it append-only makes a hand-read of an old file far less confusing.
#   target_bar   Quickfixwick's exit: one tick past the entry bar's own wick.
#   exit_close   a market-on-close, on whichever bar the strategy names.
#   exit_open    a market-on-open, on whichever bar the strategy names.
# The last two name the ORDER TYPE, not the bar, so all five bar-exit strategies share them
# (bar 0's close through bar 2's close) and none of them needed a new reason. That is
# deliberate: the reason is what the ledger says HAPPENED, and "sold at the close" is the
# same event whether it was bar 0's close or bar 2's.
# A new exit reason has to be added HERE, to `prettyReason` in build_equity_html.py AND to
# charter's TRADE_COLORS, or it will unpack as undefined on the pages and draw grey on the
# charts.
VAR_REASONS = ["target_r", "stop", "unknown_pl", "bullish_reversal", "bearish_reversal",
               "data_end", "open_at_end", "target_bar", "exit_close", "exit_open"]


def _pack_rows(raw, mkt_ix, day_ix):
    """One cap's trades as arrays of small integers and numbers, in VAR_COLS order.

    75 settings x ~80 trades of named JSON fields would put most of a megabyte of key names into
    every page. Indexing the market names and the dates against tables the page already
    has, and dropping to positional arrays, cuts that by about three quarters. Net R is not
    stored: it is gr - cr, and the page computes it.

    Three columns are here only so the PAGE can re-run the whole money management itself,
    at any risk %, without a rebuild:

      xd  the day the trade RELEASES its risk. Normally the exit date, but an open-at-end
          trade releases on its market's last bar, which is not shown anywhere else.
      gr  gross R (before slippage) -- needed to replay the gross curve alongside the net one.
      cr  cost R (the slippage charge). Both are risk-INDEPENDENT, which is exactly why the
          risk replay works: only the dollar sizing changes with risk, never the R multiples.
          (The CAP is a different matter -- it changes gr itself, which is why every cap is
          backtested here instead of being replayed.)

    PRECISION MATTERS HERE. gr and cr carry 6 decimals, not the 3 the table displays: the
    replay compounds them across ~80 trades, and rounding R to 3 dp put the page $13.56
    away from the server's own final figure (measured 2026-07-25). The display formats to
    2 dp regardless, so the extra digits cost nothing visible.
    """
    rows = []
    for t in sorted(raw, key=lambda x: (x["entry_d"], x["market"])):
        reason = t["exit_reason"]
        pout = t["stop"] if reason == "unknown_pl" else t.get("exit_price")
        rows.append([
            mkt_ix[t["market"]],
            0 if t["side"] == "short" else 1,
            day_ix[t["entry_date"]],
            None if reason == "open_at_end" else day_ix[t["exit_date"]],
            day_ix[str(t["exit_d"])],
            round(t["gross_r"], 6), round(t["cost_r"], 6),
            t["bars_in_trade"],
            t["entry"], pout, t["stop"],
            VAR_REASONS.index(reason),
        ])
    return rows


def write_variants(results, caps=None, extra=None):
    """Backtest results for every Rule 4 setting, packed into output/_variants.json.

    `caps` is the dial's axis (strategies.CAP_CHOICES by default). `extra` is every
    registered Rule 4 that is NOT a cap, defaulting to exactly those in the registry -- they
    go into the same packed table so the pages have one unpacker, and are listed under
    `extra` in the document so the dial and the levered chart, which are about the cap axis,
    can ignore them.

    The whole grid shares ONE day calendar, so moving the dial does not shift the equity
    curve's x-axis underneath the reader -- 4R and 8R are drawn over exactly the same
    period. The calendar starts at the earliest first entry across all settings; settings
    whose own first trade is later simply open flat.
    """
    caps = list(strategies.CAP_CHOICES) if caps is None else list(caps)
    if extra is None:
        extra = [s.r4 for s in strategies.REGISTRY if not s.r4.in_grid]
    # One uniform list of Rule 4 objects: the packing, the money-management replay and the
    # per-setting texts are identical whether or not the setting is a point on the dial.
    grid = [strategies.cap_rule4(c) for c in caps]
    seen = {r.token for r in grid}
    variants = grid + [r for r in extra if r.token not in seen]

    universe = [m["name"] for m in results]
    mkt_ix = {name: i for i, name in enumerate(universe)}

    prepared = {}
    for r4 in variants:
        tok = r4.token
        raw, all_days = collect(results, lambda m, tok=tok: m["var"][tok]["trades"])
        prepared[tok] = (prepare(raw), all_days)

    shared_first = min(first_trading_day(raw, days) for raw, days in prepared.values())
    all_days = next(iter(prepared.values()))[1]
    days = [d for d in all_days if d >= shared_first]
    day_ix = {str(d): i for i, d in enumerate(days)}

    out = {}
    for r4 in variants:
        raw, _ = prepared[r4.token]
        # Explicitly the REFERENCE risk. The grid is shared by every page, so it cannot be
        # priced at any one strategy's own risk -- and the pages self-check their replay
        # against these figures at exactly this number.
        with at_risk(eng.RISK_PCT):
            _, stats = account(raw, all_days, shared_first)
        out[r4.token] = dict(
            rows=_pack_rows(raw, mkt_ix, day_ix),
            # the page self-checks its own replay against this at the default risk
            final=round(stats["final"], 2),
            avg_win_r=round(stats["avg_win_r"], 3), best_r=round(stats["best_r"], 3),
            avg_bars=round(stats["avg_bars"], 2))

    # TWO lists over one `v` table. `caps` is the single cap axis -- the dial's and both
    # charts' -- and `extra` is the settings that are not caps at all. Guarded by membership
    # in `out` so a partial run (export_charter_trades asks for four caps) cannot name a
    # token it has no rows for.
    have = lambda cs: [t for t in (strategies.cap_token(c) for c in cs) if t in out]
    doc = dict(caps=have([None] + strategies.CAP_GRID),
               extra=[r.token for r in variants if not r.in_grid],
               labels={r.token: r.label for r in variants},
               texts={r.token: r.texts for r in variants},
               cols=VAR_COLS, reasons=VAR_REASONS,
               markets=universe, days=[str(d) for d in days],
               defaults={s.key: s.token for s in strategies.REGISTRY},
               v=out)
    VARIANTS_PATH.parent.mkdir(parents=True, exist_ok=True)
    VARIANTS_PATH.write_text(json.dumps(doc, separators=(",", ":")), encoding="utf-8")
    size = VARIANTS_PATH.stat().st_size
    print(f"  written: {VARIANTS_PATH.name}  {len(grid)} caps"
          + (f" + {len(variants) - len(grid)} other" if len(variants) > len(grid) else "")
          + f"  {size:,} bytes")


def main(argv):
    picked = strategies.selected(argv)
    results = eng.run_markets(picked)
    print()
    for s in picked:
        run(s, results)
    write_variants(results)


def _replay(raw, all_days, first_day, rkey):
    """Replay the same money-management with an alternative R key; return final cash only."""
    ent, ex = defaultdict(list), defaultdict(list)
    for i, t in enumerate(raw):
        ent[t["entry_d"]].append(i)
        ex[t["exit_d"]].append(i)
    cash, committed, open_risk = START_CAP, 0.0, {}
    for day in all_days:
        if first_day and day < first_day:
            continue
        for i in sorted(ent.get(day, []), key=lambda j: raw[j]["market"]):
            risk_d = max(cash - committed, 0.0) * RISK_PCT / 100.0
            open_risk[i] = risk_d
            committed += risk_d
        for i in ex.get(day, []):
            risk_d = open_risk.pop(i)
            committed -= risk_d
            cash += raw[i][rkey] * risk_d
    return cash


def _avg_bars(closed):
    """Average holding time in bars -- the number that separates a quick from a slow exit."""
    vals = [t["bars_in_trade"] for t in closed if t.get("bars_in_trade") is not None]
    return sum(vals) / len(vals) if vals else 0.0


def _streaks_and_avgs(raw):
    """Longest win/loss run (in ENTRY order) and average win/loss in $ and %.

    % is each trade's P&L over the capital it was sized against (its liquid base at
    entry) -- it is the R multiple times RISK_PCT, near enough. Open-at-end
    trades are excluded (no realised outcome).

    ENTRY ORDER, not exit order (user, 2026-07-29). The streaks ran in exit order until
    then, on the reasoning that the account balance moves as trades CLOSE. It was wrong for
    the reader: the blotter is sorted by entry, so somebody counting losing rows down the
    page counts the run in entry order, and the report was quoting a different sequence. It
    showed up on quickfixwick, where three losing positions were opened on 2026-01-29 and
    one of them did not close until 2026-02-03, after an unrelated winner had closed in
    between -- so a run the reader plainly sees as 5 was reported as 4.
    The question this number answers is "how many positions in a row lost", which is about
    the order they were TAKEN. The drawdown, which is about the order they closed, is a
    separate figure and still measured in exit order by `account()`.
    Ties inside a day are broken by market name, the same order everything else sizes in.

    It keys off net R rather than dollars, like the win rate, so it does not depend on the
    bet size: whether a position won is not a function of how big it was. That also keeps it
    meaningful at risk = 0, where every P&L in dollars is 0 and the old test scored every
    trade as neither a win nor a loss.
    """
    closed = [t for t in raw if t["exit_reason"] != "open_at_end"]
    seq = sorted(closed, key=lambda t: (t["entry_d"], t["market"]))
    lw = ll = cw = cl = 0
    for t in seq:
        p = t["r"]
        cw, cl = (cw + 1, 0) if p > 0 else ((0, cl + 1) if p < 0 else (0, 0))
        lw, ll = max(lw, cw), max(ll, cl)
    wins = [t for t in closed if t["pnl_dollars"] > 0]
    loss = [t for t in closed if t["pnl_dollars"] < 0]
    pct = lambda t: (t["pnl_dollars"] / t["base_at_entry"] * 100.0) if t.get("base_at_entry") else 0.0
    mean = lambda xs, f: (sum(f(x) for x in xs) / len(xs)) if xs else 0.0
    best = max((t["r"] for t in closed), default=0.0)
    return dict(long_win=lw, long_loss=ll, best_r=best,
                avg_win=mean(wins, lambda t: t["pnl_dollars"]),
                avg_loss=mean(loss, lambda t: t["pnl_dollars"]),
                avg_win_pct=mean(wins, pct), avg_loss_pct=mean(loss, pct),
                avg_win_r=mean(wins, lambda t: t["r"]))


def _write_json(strategy, raw, timeline, stats, universe):
    """Per-trade rows and headline stats for the equity-curve page, at the default Rule 4.

    `universe` is every market backtested, with its obsolete flag -- the page seeds the
    per-market table from it so markets that produced no trade still show, as zeros.

    No daily series here any more: the page rebuilds it from the trades on every risk AND
    every cap change, and the day calendar it walks is shared by the whole cap grid, so it
    lives in _variants.json instead of being repeated per strategy.
    """
    days = [str(t["date"]) for t in timeline]
    out = dict(strategy=strategy.key, title=strategy.title, rule4=strategy.rule4,
               # the variant token this page opens at. Named r4, not cap: it is a cap token
               # for the family and something else entirely for a strategy outside it.
               r4=strategy.token,
               risk_pct=strategy.risk_pct,   # this strategy's own default; the page opens here
               # the same account at the reference risk, for the grid-vs-workbook guard
               ref_final=round(stats["ref_final"], 2),
               start=START_CAP,
               final=round(stats["final"], 2),
               ret=round(stats["ret"] / 100.0, 4), maxdd=round(-stats["max_dd"], 2),
               first=days[0], last=days[-1], n=len(days),
               n_trades=stats["n_trades"], win_rate=stats["win_rate"],
               n_markets=stats["n_markets"], n_markets_all=stats["n_markets_all"],
               markets_all=universe,
               max_open=stats["max_open"], time_in_market=round(stats["time_in_market"], 1),
               gross_final=round(stats["gross_final"], 2),
               gross_ret=round(stats["gross_ret"] / 100.0, 4),
               total_cost=round(stats["total_cost"], 2),
               long_win=stats["long_win"], long_loss=stats["long_loss"],
               avg_win=round(stats["avg_win"], 2), avg_loss=round(stats["avg_loss"], 2),
               avg_win_pct=round(stats["avg_win_pct"], 2),
               avg_loss_pct=round(stats["avg_loss_pct"], 2),
               avg_win_r=round(stats["avg_win_r"], 3), best_r=round(stats["best_r"], 3),
               avg_bars=round(stats["avg_bars"], 2),
               slippage=dict(entry=SLIP_ENTRY_TICKS, target=SLIP_TARGET_TICKS,
                             stop=SLIP_STOP_TICKS))
    data_json(strategy).write_text(json.dumps(out, separators=(",", ":")), encoding="utf-8")


# --- workbook ---------------------------------------------------------------------
HEAD_FILL = PatternFill("solid", fgColor="1F3864")
HEAD_FONT = Font(bold=True, color="FFFFFF")
POS, NEG = Font(color="1B7A34"), Font(color="B02418")
TL_COLS = ["date", "cash_balance", "committed_risk", "liquid_capital", "open_positions",
           "open_markets", "entries_today", "exits_today"]
TR_COLS = ["market", "side", "entry_date", "exit_date", "bars_in_trade",
           "gross_r", "cost_r", "net_r", "base_at_entry", "risk_$", "pnl_$",
           "balance_after", "exit_reason"]


def _head(ws, cols):
    for c, name in enumerate(cols, 1):
        cell = ws.cell(1, c, name)
        cell.fill, cell.font = HEAD_FILL, HEAD_FONT
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"


def _autosize(ws, cols, cap=48):
    for c in range(1, len(cols) + 1):
        w = max((len(str(ws.cell(r, c).value or "")) for r in range(1, ws.max_row + 1)),
                default=8)
        ws.column_dimensions[get_column_letter(c)].width = min(w + 2, cap)


def _write(strategy, raw, timeline, stats):
    wb = Workbook()
    _sheet_summary(wb.active, strategy, stats)
    _sheet_equity(wb.create_sheet("equity_curve"), strategy, timeline)
    _sheet_trades(wb.create_sheet("trades"), raw)
    path = out_xlsx(strategy)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _sheet_summary(ws, strategy, s):
    ws.title = "summary"
    ws["A1"] = f"{strategy.key} portfolio -- one shared account, all markets"
    ws["A1"].font = Font(bold=True, size=13)
    rows = [
        ("rule 4", strategy.rule4),
        ("starting capital", f"{START_CAP:,.0f}"),
        ("final capital (net)", f"{s['final']:,.2f}"),
        ("total return (net)", f"{s['ret']:+.2f}%"),
        ("final capital (gross)", f"{s['gross_final']:,.0f}  ({s['gross_ret']:+.2f}%)"),
        ("slippage paid", f"{s['total_cost']:,.0f}  ({s['gross_ret'] - s['ret']:.1f} pts of return)"),
        ("max drawdown", f"{s['max_dd']:.2f}%"),
        # Return per point of drawdown: the ranking metric between strategies. Risk per trade
        # is a dial, so a shallower edge can be levered up to a given drawdown -- a bigger
        # total return bought with a deeper hole is not automatically the better strategy.
        ("return / drawdown",
         f"{s['ret'] / s['max_dd']:.1f}x" if s["max_dd"] > 0 else "n/a"),
        ("closed trades", s["n_trades"]),
        ("win rate", f"{s['win_rate']:.1f}%"),
        ("average hold", f"{s['avg_bars']:.1f} bars"),
        ("average winner", f"{s['avg_win_r']:+.2f}R  (best {s['best_r']:+.2f}R)"),
        ("longest win streak", s["long_win"]),
        ("longest loss streak", s["long_loss"]),
        ("average win", f"{s['avg_win']:,.0f}  ({s['avg_win_pct']:+.2f}%)"),
        ("average loss", f"{s['avg_loss']:,.0f}  ({s['avg_loss_pct']:+.2f}%)"),
        ("markets tested", s["n_markets_all"]),
        ("markets with trades", s["n_markets"]),
        ("max concurrent positions", s["max_open"]),
        ("time in market", f"{s['time_in_market']:.0f}% of trading days"),
        ("peak risk committed", f"{s['max_committed_pct']:.1f}% of live balance"),
        ("window", f"{s['first']} -> {s['last']}"),
        ("risk per trade",
         f"{strategy.risk_pct:g}% of liquid capital "
         f"(solved for a {eng.TARGET_DD:g}% max drawdown)"),
        ("slippage model",
         f"entry {SLIP_ENTRY_TICKS}t, target {SLIP_TARGET_TICKS}t, stop {SLIP_STOP_TICKS}t"),
    ]
    for i, (k, v) in enumerate(rows, 3):
        ws.cell(i, 1, k).font = Font(bold=True)
        ws.cell(i, 2, v)
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 46


def _sheet_equity(ws, strategy, timeline):
    _head(ws, TL_COLS)
    for t in timeline:
        ws.append([t["date"], round(t["cash"], 2), round(t["committed"], 2),
                   round(t["liquid"], 2), t["open_count"], t["open_markets"],
                   t["entries"], t["exits"]])
        ws.cell(ws.max_row, 1).number_format = "yyyy-mm-dd"
        for c in (2, 3, 4):
            ws.cell(ws.max_row, c).number_format = "#,##0"
    _autosize(ws, TL_COLS)
    # equity chart: cash balance vs date
    chart = LineChart()
    chart.title = f"{strategy.title} portfolio capital (realized cash balance)"
    chart.y_axis.title, chart.x_axis.title = "capital", "date"
    chart.height, chart.width = 9, 32
    last = ws.max_row
    chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=last), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=last))
    chart.legend = None
    ws.add_chart(chart, "J2")


def _sheet_trades(ws, raw):
    _head(ws, TR_COLS)
    for t in sorted(raw, key=lambda x: (x["entry_d"], x["market"])):
        pnl = t.get("pnl_dollars")
        ws.append([t["market"], t["side"], t["entry_date"],
                   t["exit_date"] if t["exit_reason"] != "open_at_end" else None,
                   t["bars_in_trade"], round(t["gross_r"], 3), round(t["cost_r"], 3),
                   round(t["r"], 3),
                   round(t.get("base_at_entry", 0), 2), round(t.get("risk_dollars", 0), 2),
                   round(pnl, 2) if pnl is not None else None,
                   round(t["balance_after"], 2) if t.get("balance_after") is not None else None,
                   t["exit_reason"]])
        r = ws.max_row
        for c in (9, 10, 11, 12):
            ws.cell(r, c).number_format = "#,##0"
        if pnl is not None:
            pc = ws.cell(r, TR_COLS.index("pnl_$") + 1)
            pc.font = POS if pnl > 0 else (NEG if pnl < 0 else Font())
    _autosize(ws, TR_COLS)


if __name__ == "__main__":
    main(sys.argv[1:])
