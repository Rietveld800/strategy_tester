# quickfix_portfolio.py
#
# Combine every market's quickfix trades into ONE shared account and build a chronological
# equity curve of the capital. The per-market trades themselves are unchanged (entries,
# exits and R-multiples depend only on price/reversals, not capital); only the money
# management is re-run on the shared account.
#
# Money management (confirmed with the user):
#   - one account, starting STARTING_CAPITAL; cash balance changes ONLY when a trade closes.
#   - a new trade risks 1% of LIQUID capital = cash balance - risk already tied up in the
#     trades currently open. Each open trade ties up its own 1% until it closes.
#   - no cap on how many trades are open at once.
#   - within a date: process ALL entries first (sized while that day's closing trades still
#     tie up their risk), THEN the exits (which book P&L and grow the balance).
#   - several entries on one date are sized in a fixed order (alphabetical by market), each
#     off the base left by the earlier ones.
#
# An "open at end" position (data ended while it was open) ties up its 1% from entry until
# its market's last bar, then releases it with zero realised P&L -- its outcome is unknown.

import json
from collections import defaultdict
from datetime import date

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import quickfix as qf
from quickfix_all import market_dirs

OUT_XLSX = qf.HERE / "output" / "quickfix_portfolio_daily.xlsx"
DATA_JSON = qf.HERE / "output" / "_equity_data.json"
RISK_PCT = qf.RISK_PCT
START_CAP = qf.STARTING_CAPITAL

# --- transaction costs ------------------------------------------------------------
# Realistic fills, charged as SLIPPAGE in ticks (the market's inferred price tick, the
# same unit the stop uses). Entry and the limit take-profit fill at/near their price; a
# stop is a market order on trigger and slips more, especially on gaps. Cost converts to
# R via each trade's own risk distance -- ticks * tick / risk_per_unit -- so a wide-risk
# trade barely feels it and a tight-risk trade feels it more, which is how slippage really
# behaves. In dollars this equals ticks * tick * position size, the true slippage cost.
SLIP_ENTRY_TICKS = 1       # entry fill one tick worse than the reversal price
SLIP_TARGET_TICKS = 1      # limit take-profit / reversal exit
SLIP_STOP_TICKS = 3        # stop-loss and unknown-outcome exits slip more (market/gap)


def cost_in_r(trade):
    """Round-trip slippage for one trade, expressed in R (fraction of the 1% risked)."""
    rpu = trade.get("risk_per_unit")
    if not rpu or rpu <= 0:
        return 0.0
    reason = trade["exit_reason"]
    if reason == "open_at_end":
        exit_ticks = 0                                   # never exited -> no exit slippage
    elif reason in ("stop", "unknown_pl"):
        exit_ticks = SLIP_STOP_TICKS
    else:
        exit_ticks = SLIP_TARGET_TICKS
    return (SLIP_ENTRY_TICKS + exit_ticks) * trade["tick"] / rpu


def collect():
    """Every market's trades tagged with market + last bar, plus the set of all trading days."""
    raw, all_days = [], set()
    for d in market_dirs():
        bars = qf.load_bars(d)
        if not bars:
            continue
        tick, dp = qf.infer_tick(bars)
        res = qf.backtest(bars, tick, dp)
        last_bar = bars[-1].date.date()
        all_days.update(b.date.date() for b in bars)
        for t in res["trades"]:
            raw.append({**t, "market": d.name, "last_bar": last_bar, "tick": tick})
    return raw, sorted(all_days)


def run_portfolio():
    raw, all_days = collect()

    # normalise each trade to (entry day, exit day, gross R, cost, net R). An open-at-end
    # trade "exits" on its market's last bar with gross R = 0 (outcome unknown).
    for t in raw:
        t["entry_d"] = date.fromisoformat(t["entry_date"])
        if t["exit_reason"] == "open_at_end":
            t["exit_d"], t["gross_r"] = t["last_bar"], 0.0
        else:
            t["exit_d"], t["gross_r"] = date.fromisoformat(t["exit_date"]), t["r_multiple"]
        t["cost_r"] = cost_in_r(t)
        t["r"] = t["gross_r"] - t["cost_r"]          # net R actually booked

    entries_by_day, exits_by_day = defaultdict(list), defaultdict(list)
    for i, t in enumerate(raw):
        entries_by_day[t["entry_d"]].append(i)
        exits_by_day[t["exit_d"]].append(i)

    cash, committed = START_CAP, 0.0
    open_risk = {}                         # trade index -> risk dollars tied up
    timeline, peak, max_dd = [], START_CAP, 0.0
    max_open, max_committed_ratio, total_cost, days_in_mkt = 0, 0.0, 0.0, 0
    first_day = min(t["entry_d"] for t in raw) if raw else (all_days[0] if all_days else None)

    for day in all_days:
        if first_day and day < first_day:
            continue
        # 1) ENTRIES first -- sized while that day's closing trades still tie up their risk
        ent_notes = []
        for i in sorted(entries_by_day.get(day, []), key=lambda j: raw[j]["market"]):
            base = max(cash - committed, 0.0)        # guard: never size on negative capital
            risk_d = base * RISK_PCT / 100.0
            raw[i]["risk_dollars"] = risk_d
            raw[i]["base_at_entry"] = base
            open_risk[i] = risk_d
            committed += risk_d
            ent_notes.append(f"{raw[i]['market']} {raw[i]['side']} (risk {risk_d:,.0f})")

        # 2) EXITS -- book P&L and release the tied-up risk
        ex_notes = []
        for i in exits_by_day.get(day, []):
            risk_d = open_risk.pop(i)
            committed -= risk_d
            pnl = raw[i]["r"] * risk_d
            cash += pnl
            total_cost += raw[i]["cost_r"] * risk_d       # slippage paid, in dollars
            raw[i]["pnl_dollars"] = pnl
            raw[i]["balance_after"] = cash
            tag = "open->closed 0" if raw[i]["exit_reason"] == "open_at_end" else f"{pnl:+,.0f}"
            ex_notes.append(f"{raw[i]['market']} {raw[i]['exit_reason']} ({tag})")

        liquid = cash - committed
        peak = max(peak, cash)
        dd = (cash - peak) / peak * 100.0
        max_dd = max(max_dd, -dd)
        max_open = max(max_open, len(open_risk))
        if len(open_risk) > 0:
            days_in_mkt += 1
        if cash > 0:
            max_committed_ratio = max(max_committed_ratio, committed / cash * 100.0)
        timeline.append(dict(
            date=day, cash=cash, committed=committed, liquid=liquid, dd=dd,
            open_count=len(open_risk),
            open_markets=", ".join(sorted(raw[i]["market"] for i in open_risk)),
            entries="; ".join(ent_notes), exits="; ".join(ex_notes)))

    gross_final = _replay(raw, all_days, first_day, "gross_r")   # same trades, no costs
    closed = [t for t in raw if t["exit_reason"] != "open_at_end"]
    wins = sum(1 for t in closed if t["r"] > 0)
    n_days = len(timeline)
    stats = dict(final=cash, ret=(cash / START_CAP - 1) * 100.0, max_dd=max_dd,
                 n_trades=len(closed), wins=wins,
                 win_rate=100 * wins / len(closed) if closed else 0.0,
                 max_open=max_open, max_committed_pct=max_committed_ratio,
                 first=first_day, last=all_days[-1] if all_days else None,
                 n_markets=len({t["market"] for t in raw}),
                 gross_final=gross_final, gross_ret=(gross_final / START_CAP - 1) * 100.0,
                 total_cost=total_cost,
                 time_in_market=100 * days_in_mkt / n_days if n_days else 0.0)
    stats.update(_streaks_and_avgs(raw))
    _write(raw, timeline, stats)
    _write_json(raw, timeline, stats)
    print(f"NET  ${cash:,.2f} ({stats['ret']:+.2f}%)   gross ${gross_final:,.0f} "
          f"({stats['gross_ret']:+.2f}%)   cost drag ${total_cost:,.0f}")
    print(f"maxDD {max_dd:.2f}%  trades {len(closed)} winrate {stats['win_rate']:.1f}%  "
          f"max concurrent {max_open}  time in market {stats['time_in_market']:.0f}%")
    print(f"written: {OUT_XLSX}")


def _replay(raw, all_days, first_day, rkey):
    """Replay the same money-management with an alternative R key; return final cash only."""
    ent, ex = defaultdict(list), defaultdict(list)
    for i, t in enumerate(raw):
        ent[t["entry_d"]].append(i)
        ex[t["exit_d"]].append(i)
    cash, committed, open_risk = START_CAP, 0.0, {}
    for day in all_days:
        if first_day and day < first_day:
            continue
        for i in sorted(ent.get(day, []), key=lambda j: raw[j]["market"]):
            risk_d = max(cash - committed, 0.0) * RISK_PCT / 100.0
            open_risk[i] = risk_d
            committed += risk_d
        for i in ex.get(day, []):
            risk_d = open_risk.pop(i)
            committed -= risk_d
            cash += raw[i][rkey] * risk_d
    return cash


def _streaks_and_avgs(raw):
    """Longest win/loss run (in exit order) and average win/loss in $ and %.

    % is each trade's P&L over the capital it was sized against (its liquid base at
    entry) -- with 1% risk this is essentially the R multiple as a percent. Open-at-end
    trades are excluded (no realised outcome).
    """
    closed = [t for t in raw if t["exit_reason"] != "open_at_end"]
    seq = sorted(closed, key=lambda t: (t["exit_d"], t["entry_d"], t["market"]))
    lw = ll = cw = cl = 0
    for t in seq:
        p = t["pnl_dollars"]
        cw, cl = (cw + 1, 0) if p > 0 else ((0, cl + 1) if p < 0 else (0, 0))
        lw, ll = max(lw, cw), max(ll, cl)
    wins = [t for t in closed if t["pnl_dollars"] > 0]
    loss = [t for t in closed if t["pnl_dollars"] < 0]
    pct = lambda t: (t["pnl_dollars"] / t["base_at_entry"] * 100.0) if t.get("base_at_entry") else 0.0
    mean = lambda xs, f: (sum(f(x) for x in xs) / len(xs)) if xs else 0.0
    return dict(long_win=lw, long_loss=ll,
                avg_win=mean(wins, lambda t: t["pnl_dollars"]),
                avg_loss=mean(loss, lambda t: t["pnl_dollars"]),
                avg_win_pct=mean(wins, pct), avg_loss_pct=mean(loss, pct))


def _trade_rows(raw):
    """Per-trade records for the sortable table on the page."""
    rows = []
    for t in sorted(raw, key=lambda x: (x["entry_d"], x["market"])):
        b, pnl = t.get("base_at_entry", 0), t.get("pnl_dollars", 0.0)
        reason = t["exit_reason"]
        pout = t["stop"] if reason == "unknown_pl" else t.get("exit_price")
        rows.append(dict(
            market=t["market"], side=t["side"], din=t["entry_date"],
            dout=(t["exit_date"] if reason != "open_at_end" else None),
            bars=t["bars_in_trade"], pin=t["entry"], pout=pout, sl=t["stop"],
            tgt=t["target_5r"], r=round(t["r"], 3),
            pnl=round(pnl, 2), pnlpct=round(pnl / b * 100.0, 3) if b else 0.0,
            reason=reason))
    return rows


def _write_json(raw, timeline, stats):
    """Compact daily series, per-trade rows, and headline stats for the equity-curve page."""
    pts = [dict(date=str(t["date"]), cash=round(t["cash"], 2), open=t["open_count"],
                markets=t["open_markets"], dd=round(t["dd"], 2),
                entries=t["entries"], exits=t["exits"]) for t in timeline]
    out = dict(points=pts, trades=_trade_rows(raw), start=START_CAP,
               final=round(stats["final"], 2),
               ret=round(stats["ret"] / 100.0, 4), maxdd=round(-stats["max_dd"], 2),
               first=pts[0]["date"], last=pts[-1]["date"], n=len(pts),
               n_trades=stats["n_trades"], win_rate=stats["win_rate"],
               max_open=stats["max_open"], time_in_market=round(stats["time_in_market"], 1),
               gross_final=round(stats["gross_final"], 2),
               gross_ret=round(stats["gross_ret"] / 100.0, 4),
               total_cost=round(stats["total_cost"], 2),
               long_win=stats["long_win"], long_loss=stats["long_loss"],
               avg_win=round(stats["avg_win"], 2), avg_loss=round(stats["avg_loss"], 2),
               avg_win_pct=round(stats["avg_win_pct"], 2),
               avg_loss_pct=round(stats["avg_loss_pct"], 2))
    DATA_JSON.write_text(json.dumps(out, separators=(",", ":")), encoding="utf-8")


# --- workbook ---------------------------------------------------------------------
HEAD_FILL = PatternFill("solid", fgColor="1F3864")
HEAD_FONT = Font(bold=True, color="FFFFFF")
POS, NEG = Font(color="1B7A34"), Font(color="B02418")
TL_COLS = ["date", "cash_balance", "committed_risk", "liquid_capital", "open_positions",
           "open_markets", "entries_today", "exits_today"]
TR_COLS = ["market", "side", "entry_date", "exit_date", "bars_in_trade",
           "gross_r", "cost_r", "net_r", "base_at_entry", "risk_$", "pnl_$",
           "balance_after", "exit_reason"]


def _head(ws, cols):
    for c, name in enumerate(cols, 1):
        cell = ws.cell(1, c, name)
        cell.fill, cell.font = HEAD_FILL, HEAD_FONT
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"


def _autosize(ws, cols, cap=48):
    for c in range(1, len(cols) + 1):
        w = max((len(str(ws.cell(r, c).value or "")) for r in range(1, ws.max_row + 1)),
                default=8)
        ws.column_dimensions[get_column_letter(c)].width = min(w + 2, cap)


def _write(raw, timeline, stats):
    wb = Workbook()
    _sheet_summary(wb.active, stats)
    _sheet_equity(wb.create_sheet("equity_curve"), timeline)
    _sheet_trades(wb.create_sheet("trades"), raw)
    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)


def _sheet_summary(ws, s):
    ws.title = "summary"
    ws["A1"] = "quickfix portfolio -- one shared account, all markets"
    ws["A1"].font = Font(bold=True, size=13)
    rows = [
        ("starting capital", f"{START_CAP:,.0f}"),
        ("final capital (net)", f"{s['final']:,.2f}"),
        ("total return (net)", f"{s['ret']:+.2f}%"),
        ("final capital (gross)", f"{s['gross_final']:,.0f}  ({s['gross_ret']:+.2f}%)"),
        ("slippage paid", f"{s['total_cost']:,.0f}  ({s['gross_ret'] - s['ret']:.1f} pts of return)"),
        ("max drawdown", f"{s['max_dd']:.2f}%"),
        ("closed trades", s["n_trades"]),
        ("win rate", f"{s['win_rate']:.1f}%"),
        ("longest win streak", s["long_win"]),
        ("longest loss streak", s["long_loss"]),
        ("average win", f"{s['avg_win']:,.0f}  ({s['avg_win_pct']:+.2f}%)"),
        ("average loss", f"{s['avg_loss']:,.0f}  ({s['avg_loss_pct']:+.2f}%)"),
        ("markets with trades", s["n_markets"]),
        ("max concurrent positions", s["max_open"]),
        ("time in market", f"{s['time_in_market']:.0f}% of trading days"),
        ("peak risk committed", f"{s['max_committed_pct']:.1f}% of live balance"),
        ("window", f"{s['first']} -> {s['last']}"),
        ("risk per trade", f"{RISK_PCT}% of liquid capital"),
        ("slippage model",
         f"entry {SLIP_ENTRY_TICKS}t, target {SLIP_TARGET_TICKS}t, stop {SLIP_STOP_TICKS}t"),
    ]
    for i, (k, v) in enumerate(rows, 3):
        ws.cell(i, 1, k).font = Font(bold=True)
        ws.cell(i, 2, v)
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 30


def _sheet_equity(ws, timeline):
    _head(ws, TL_COLS)
    for t in timeline:
        ws.append([t["date"], round(t["cash"], 2), round(t["committed"], 2),
                   round(t["liquid"], 2), t["open_count"], t["open_markets"],
                   t["entries"], t["exits"]])
        ws.cell(ws.max_row, 1).number_format = "yyyy-mm-dd"
        for c in (2, 3, 4):
            ws.cell(ws.max_row, c).number_format = "#,##0"
    _autosize(ws, TL_COLS)
    # equity chart: cash balance vs date
    chart = LineChart()
    chart.title = "Portfolio capital (realized cash balance)"
    chart.y_axis.title, chart.x_axis.title = "capital", "date"
    chart.height, chart.width = 9, 32
    last = ws.max_row
    chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=last), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=last))
    chart.legend = None
    ws.add_chart(chart, "J2")


def _sheet_trades(ws, raw):
    _head(ws, TR_COLS)
    for t in sorted(raw, key=lambda x: (x["entry_d"], x["market"])):
        pnl = t.get("pnl_dollars")
        ws.append([t["market"], t["side"], t["entry_date"],
                   t["exit_date"] if t["exit_reason"] != "open_at_end" else None,
                   t["bars_in_trade"], round(t["gross_r"], 3), round(t["cost_r"], 3),
                   round(t["r"], 3),
                   round(t.get("base_at_entry", 0), 2), round(t.get("risk_dollars", 0), 2),
                   round(pnl, 2) if pnl is not None else None,
                   round(t["balance_after"], 2) if t.get("balance_after") is not None else None,
                   t["exit_reason"]])
        r = ws.max_row
        for c in (9, 10, 11, 12):
            ws.cell(r, c).number_format = "#,##0"
        if pnl is not None:
            pc = ws.cell(r, TR_COLS.index("pnl_$") + 1)
            pc.font = POS if pnl > 0 else (NEG if pnl < 0 else Font())
    _autosize(ws, TR_COLS)


if __name__ == "__main__":
    run_portfolio()
