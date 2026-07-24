# quickfix_all.py
#
# Run the quickfix strategy (see quickfix.py) over EVERY market's daily array files and
# write one xlsx workbook with the results. Obsolete markets (data collection stopped) are
# included so their trades can be reviewed too.
#
# Each market is an INDEPENDENT backtest starting from a fresh STARTING_CAPITAL, so the
# per-market returns are directly comparable. The window is data-driven: trading begins
# once a market's reversals are first reported (older files carry no reversal block) and
# runs to its last daily bar.

from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import quickfix as qf

OBSOLETE_AFTER_DAYS = 30          # charter's rule: lagging the newest daily file by > 30d
OUT_XLSX = qf.HERE / "output" / "quickfix_all_markets_daily.xlsx"

TRADE_COLS = ["market", "id", "side", "entry_date", "exit_date", "bars_in_trade",
              "entry", "stop", "target_5r", "exit_price", "exit_reason",
              "r_multiple", "pnl_pct", "equity_before", "equity_after"]
SUMMARY_COLS = ["market", "obsolete", "first_reversal", "last_bar", "bars", "trades",
                "wins", "stops", "unknown", "open_at_end", "win_rate_%",
                "return_%", "final_equity"]


def market_dirs():
    """Every market directory that has daily array files (skip #Charts, _vps2, etc.)."""
    root = qf.ARRAY_ROOT
    out = []
    for d in sorted(root.iterdir()):
        if not d.is_dir() or d.name.startswith(("#", "_")):
            continue
        if any(d.glob(f"*/{qf.TIMEFRAME}/*_array.xlsx")):
            out.append(d)
    return out


def first_reversal_date(bars):
    """First bar carrying BOTH bullish and bearish reversals -- the strategy's real start."""
    for b in bars:
        if b.bull and b.bear:
            return b.date
    return None


def run_all():
    dirs = market_dirs()
    per_market = []
    newest = None
    for d in dirs:
        bars = qf.load_bars(d)
        if not bars:
            continue
        tick, dp = qf.infer_tick(bars)
        res = qf.backtest(bars, tick, dp)
        last_bar = bars[-1].date
        newest = last_bar if newest is None else max(newest, last_bar)
        per_market.append(dict(name=d.name, bars=bars, res=res,
                               first_rev=first_reversal_date(bars), last_bar=last_bar))
        closed = [t for t in res["trades"] if t["exit_reason"] != "open_at_end"]
        print(f"{d.name:38} bars={len(bars):>4} trades={len(closed):>3} "
              f"return={ (res['final_equity']/qf.STARTING_CAPITAL-1)*100:+7.2f}%")

    # obsolescence is relative to the newest daily bar across ALL markets
    for m in per_market:
        m["obsolete"] = (newest - m["last_bar"]).days > OBSOLETE_AFTER_DAYS

    _write_xlsx(per_market, newest)
    print(f"\nwritten: {OUT_XLSX}")


def _summ_row(m):
    trades = m["res"]["trades"]
    closed = [t for t in trades if t["exit_reason"] != "open_at_end"]
    wins = sum(1 for t in closed if t["pnl_pct"] > 0)
    stops = sum(1 for t in closed if t["exit_reason"] == "stop")
    unknown = sum(1 for t in closed if t["exit_reason"] == "unknown_pl")
    open_end = sum(1 for t in trades if t["exit_reason"] == "open_at_end")
    ret = (m["res"]["final_equity"] / qf.STARTING_CAPITAL - 1) * 100
    win_rate = (100 * wins / len(closed)) if closed else None
    return dict(market=m["name"], obsolete="yes" if m["obsolete"] else "",
                first_reversal=str(m["first_rev"].date()) if m["first_rev"] else "",
                last_bar=str(m["last_bar"].date()), bars=len(m["bars"]),
                trades=len(closed), wins=wins, stops=stops, unknown=unknown,
                open_at_end=open_end,
                win_rate=round(win_rate, 1) if win_rate is not None else None,
                ret=round(ret, 2), final_equity=round(m["res"]["final_equity"], 2))


def _write_xlsx(per_market, newest):
    wb = Workbook()
    _sheet_summary(wb.active, per_market, newest)
    _sheet_trades(wb.create_sheet("trades"), per_market)
    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)


HEAD_FILL = PatternFill("solid", fgColor="1F3864")
HEAD_FONT = Font(bold=True, color="FFFFFF")
OBS_FILL = PatternFill("solid", fgColor="EEEEEE")
POS_FONT = Font(color="1B7A34")
NEG_FONT = Font(color="B02418")


def _header(ws, cols):
    for c, name in enumerate(cols, 1):
        cell = ws.cell(1, c, name)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"


def _autosize(ws, cols):
    for c, name in enumerate(cols, 1):
        width = max(len(str(name)) + 2,
                    *(len(str(ws.cell(r, c).value or "")) + 2
                      for r in range(2, ws.max_row + 1)) or [0])
        ws.column_dimensions[get_column_letter(c)].width = min(width, 42)


def _sheet_summary(ws, per_market, newest):
    ws.title = "summary"
    _header(ws, SUMMARY_COLS)
    # active markets first (by return desc), then obsolete (by return desc) -- charter's order
    rows = [_summ_row(m) for m in per_market]
    rows.sort(key=lambda r: (r["obsolete"] == "yes", -r["ret"]))
    for r in rows:
        vals = [r["market"], r["obsolete"], r["first_reversal"], r["last_bar"], r["bars"],
                r["trades"], r["wins"], r["stops"], r["unknown"], r["open_at_end"],
                r["win_rate"], r["ret"], r["final_equity"]]
        ws.append(vals)
        row = ws.max_row
        ret_cell = ws.cell(row, SUMMARY_COLS.index("return_%") + 1)
        ret_cell.number_format = "+0.00;-0.00"
        ret_cell.font = POS_FONT if r["ret"] > 0 else (NEG_FONT if r["ret"] < 0 else Font())
        ws.cell(row, SUMMARY_COLS.index("final_equity") + 1).number_format = "#,##0"
        if r["obsolete"] == "yes":
            for c in range(1, len(SUMMARY_COLS) + 1):
                ws.cell(row, c).fill = OBS_FILL
    _autosize(ws, SUMMARY_COLS)
    # footer note
    note = (f"quickfix daily, starting capital {qf.STARTING_CAPITAL:,.0f}, "
            f"risk {qf.RISK_PCT}%/trade, fees {qf.FEES}. "
            f"obsolete = last bar > {OBSOLETE_AFTER_DAYS}d before newest ({newest.date()}). "
            f"generated {datetime.now(timezone.utc):%Y-%m-%d %H:%M UTC}.")
    ws.append([])
    ws.append([note])


def _sheet_trades(ws, per_market):
    _header(ws, TRADE_COLS)
    order = sorted(per_market, key=lambda m: m["name"])
    for m in order:
        for t in m["res"]["trades"]:
            ws.append([m["name"], t["id"], t["side"], t["entry_date"], t["exit_date"],
                       t["bars_in_trade"], t["entry"], t["stop"], t["target_5r"],
                       t["exit_price"], t["exit_reason"], t["r_multiple"], t["pnl_pct"],
                       t["equity_before"], t["equity_after"]])
            row = ws.max_row
            pnl = t["pnl_pct"]
            if pnl is not None:
                cell = ws.cell(row, TRADE_COLS.index("pnl_pct") + 1)
                cell.number_format = "+0.00;-0.00"
                cell.font = POS_FONT if pnl > 0 else (NEG_FONT if pnl < 0 else Font())
    _autosize(ws, TRADE_COLS)


if __name__ == "__main__":
    run_all()
